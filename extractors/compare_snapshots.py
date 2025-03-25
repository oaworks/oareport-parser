import os
import sys
import re
import yaml
import argparse
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from collections import defaultdict

# Allows "from export.google_sheets import load_gsheet_to_df"
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from export.google_sheets import load_gsheet_to_df

##############################################
# 1. Load config and define whole-number sets
##############################################
def load_config():
    config_path = os.path.join(os.path.dirname(__file__), "..", "config", "settings.yaml")
    print(f"Loading config from: {config_path}")
    with open(config_path, "r") as f:
        return yaml.safe_load(f)

CONFIG = load_config()

INSIGHTS_WHOLE_NUMBER_METRICS = {
    "Total publications (insight)",
    "Total preprints (insight)"
}
EXPLORE_WHOLE_NUMBER_METRICS = {
    "PUBLICATIONS",
    "Total APC amount",
    "Mean APC amount",
    "Median APC amount"
}

##############################################
# 2. Basic Helpers
##############################################
def filter_by_date(df, date_str):
    """Keep only rows whose collection_time starts with YYYY-MM-DD."""
    return df[df["collection_time"].str.startswith(date_str)]

def extract_org_from_url(url):
    """Given https://dev.oa.report/gates-foundation?orgkey=abc => gates-foundation."""
    return url.split(".report/")[1].split("?")[0].strip("/")

def format_date_label(date_str):
    """'2025-03-24' => '24 Mar' (for Windows => '%#d %b', else '%-d %b')."""
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d")
        return d.strftime("%-d %b") if os.name != "nt" else d.strftime("%#d %b")
    except:
        return date_str

##############################################
# 3. Compare numeric values
##############################################
def compare_values(old, new, is_whole):
    """
    If is_whole=False => treat as percentage => /100.
    If both old & new ~ 0 => 0% (not inf).
    If old=0 & new!=0 => inf.
    Round to consistent decimals.
    """
    try:
        old_val = float(str(old).replace("%", "").replace(",", ""))
        new_val = float(str(new).replace("%", "").replace(",", ""))
        if not is_whole:
            old_val /= 100
            new_val /= 100

        delta = new_val - old_val
        if abs(old_val) < 1e-12 and abs(new_val) < 1e-12:
            pct = 0
        elif abs(old_val) < 1e-12:
            pct = float("inf")
        else:
            pct = (delta / old_val) * 100

        return round(new_val, 4), round(old_val, 4), round(delta, 4), round(pct, 2)
    except:
        return ("N/A" if pd.isna(new) else new,
                "N/A" if pd.isna(old) else old,
                "N/A",
                "N/A")

##############################################
# 4. Merging logic for Insights & Actions
##############################################
def process_rows_by_key(old_df, new_df, key_col, date_col):
    """
    Outer approach => keep all combos, 
    then compare old vs new for each (key_col, date_col).
    """
    # unify column name => 'Value'
    if "Value" not in old_df and "value" in old_df:
        old_df = old_df.rename(columns={"value": "Value"})
    if "Value" not in new_df and "value" in new_df:
        new_df = new_df.rename(columns={"value": "Value"})

    old_keys = old_df[[key_col, date_col]].drop_duplicates()
    new_keys = new_df[[key_col, date_col]].drop_duplicates()

    combos = pd.merge(old_keys, new_keys, how="outer", on=[key_col, date_col])

    results = []
    for _, row in combos.iterrows():
        metric = row[key_col]
        dr = row[date_col]

        old_part = old_df[(old_df[key_col] == metric) & (old_df[date_col] == dr)]
        new_part = new_df[(new_df[key_col] == metric) & (new_df[date_col] == dr)]

        val_old = old_part["Value"].values[0] if not old_part.empty else "N/A"
        val_new = new_part["Value"].values[0] if not new_part.empty else "N/A"

        # If it's an insight => check if it's in the WHOLE set
        if "(insight)" in str(metric) and metric in INSIGHTS_WHOLE_NUMBER_METRICS:
            is_whole = True
        elif "(insight)" in str(metric):
            is_whole = False
        else:
            # actions => all whole
            is_whole = True

        newv, oldv, delta, pct = compare_values(val_old, val_new, is_whole)
        results.append({
            "DATE_RANGE": dr if pd.notna(dr) else "N/A",
            "METRIC": metric if pd.notna(metric) else "N/A",
            "Old": oldv,
            "New": newv,
            "Change": delta,
            "% Change": pct
        })
    return pd.DataFrame(results)

##############################################
# 5. Merging logic for Explore
##############################################
def process_explore_section(old_df, new_df):
    """
    Each row => (KEY + metric).
    We'll unify columns from both old & new => partial merges with N/A if missing.
    """
    skip_cols = {"collection_time", "org_url", "KEY"}
    all_cols = list(set(old_df.columns).union(new_df.columns))
    metrics = [c for c in all_cols if c not in skip_cols]

    # unify keys (outer)
    old_keys = old_df["KEY"].drop_duplicates()
    new_keys = new_df["KEY"].drop_duplicates()
    combos = pd.merge(old_keys.to_frame("KEY"), new_keys.to_frame("KEY"), how="outer", on="KEY")

    results = []
    for dr in combos["KEY"].unique():
        old_part = old_df[old_df["KEY"] == dr]
        new_part = new_df[new_df["KEY"] == dr]

        for metric in metrics:
            val_old = old_part.iloc[0][metric] if (not old_part.empty and metric in old_part.columns) else "N/A"
            val_new = new_part.iloc[0][metric] if (not new_part.empty and metric in new_part.columns) else "N/A"

            # if metric in EXPLORE_WHOLE_NUMBER_METRICS => treat as whole
            is_whole = metric in EXPLORE_WHOLE_NUMBER_METRICS
            newv, oldv, delta, pct = compare_values(val_old, val_new, is_whole)

            results.append({
                "DATE_RANGE": dr if pd.notna(dr) else "N/A",
                "METRIC": metric,
                "Old": oldv,
                "New": newv,
                "Change": delta,
                "% Change": pct
            })
    return pd.DataFrame(results)

##############################################
# 6. Master aggregator per section
##############################################
def process_section(section, sheet_name, creds_path, date1, date2):
    df = load_gsheet_to_df(sheet_name, creds_path)
    if df.empty:
        print(f"Sheet {sheet_name} is empty! No data.")
        return {}

    # figure out org column
    url_col = next((c for c in ["org_url","Page_URL","Original_URL"] if c in df.columns), None)
    if not url_col:
        print(f"No org_url-like col for section {section}, skipping.")
        return {}

    # attach org
    df["org"] = df[url_col].apply(extract_org_from_url)

    # Filter old/new
    results_by_org = {}
    for org in df["org"].unique():
        subset = df[df["org"] == org]
        df_old = filter_by_date(subset, date1)
        df_new = filter_by_date(subset, date2)

        # if no data in either => skip
        if df_old.empty and df_new.empty:
            continue

        if section == "explore":
            out = process_explore_section(df_old.copy(), df_new.copy())
        else:
            key_col = "Insight" if section == "insights" else "strategy"
            out = process_rows_by_key(df_old.copy(), df_new.copy(), key_col=key_col, date_col="date_range")

        if not out.empty:
            results_by_org.setdefault(org, []).append(out)

    return results_by_org

##############################################
# 7. Final styling
##############################################
def apply_styling(workbook):
    """Color Red if % Change=0, Bold if |% Change|>=1, grey fill for last two cols."""
    red_font = Font(color="9C0006")
    bold_font = Font(bold=True)
    grey_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

    for ws in workbook.worksheets:
        for row in ws.iter_rows(min_row=2):
            # 'Change' => col idx=4, '% Change' => col idx=5
            change_cell = row[4]
            pct_cell = row[5]
            try:
                pct_val = float(pct_cell.value)
                if abs(pct_val) < 1e-9:
                    pct_cell.font = red_font
                elif abs(pct_val) >= 1:
                    change_cell.font = bold_font
                    pct_cell.font = bold_font
            except:
                pass
            change_cell.fill = grey_fill
            pct_cell.fill = grey_fill

##############################################
# 8. Main
##############################################
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--env", choices=["dev","staging"], required=True)
    parser.add_argument("--date1", required=True)  # e.g. '2025-03-24'
    parser.add_argument("--date2", required=True)  # e.g. '2025-03-25'
    parser.add_argument("--year", required=True)   # e.g. '2025'
    args = parser.parse_args()

    creds_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..",
        CONFIG["google_sheets"]["creds_file"]
    )
    sheets_map = CONFIG["google_sheets"]["sheets"]
    output_file = f"qa_snapshot_{args.year}.xlsx"
    writer = pd.ExcelWriter(output_file, engine="openpyxl")
    all_results = defaultdict(list)

    # Gather data from insights, actions, explore
    for section in ["insights","actions","explore"]:
        sheet_name = sheets_map[section][args.env]
        print(f"Processing {section} from {sheet_name} for {args.date1} vs {args.date2}")
        sec_result = process_section(section, sheet_name, creds_path, args.date1, args.date2)
        for org, df_list in sec_result.items():
            all_results[org].extend(df_list)

    from openpyxl import load_workbook

    # Write each org to its own sheet
    for org, dataframes in all_results.items():
        if not dataframes:
            continue
        combined = pd.concat(dataframes, ignore_index=True)

        # Insert date-labeled columns
        old_label = f"Old ({format_date_label(args.date1)})"
        new_label = f"New ({format_date_label(args.date2)})"
        combined.insert(2, old_label, combined.pop("Old"))
        combined.insert(3, new_label, combined.pop("New"))

        # Rename 'FIGURE' -> 'METRIC'
        combined.rename(columns={"FIGURE": "METRIC"}, inplace=True)

        # Sort by DATE_RANGE => METRIC
        combined.sort_values(["DATE_RANGE", "METRIC"], inplace=True, na_position="last")

        combined.to_excel(writer, sheet_name=org[:31], index=False)

    writer.close()

    # Apply styling
    wb = load_workbook(output_file)
    apply_styling(wb)
    wb.save(output_file)
    print(f"QA snapshot saved to {output_file}")

if __name__=="__main__":
    main()
